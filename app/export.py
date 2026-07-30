"""
export.py

Produces the actual output file you send to the supervisor: the exact
Excel file you uploaded, completely unchanged, except the Vehicle and
Driver columns are filled in with the finalized assignment. No other
column, formatting, header, or formula is touched -- this loads the
original workbook (which keeps all its styling) and only ever writes
into two specific cells per row.
"""
import re

import openpyxl

_HEADER_ALIASES = {
    "vehicle": "vehicle",
    "driver": "driver",
}


def _normalize_header(h):
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def export_filled_excel(original_path, jobs, output_path):
    """
    original_path: the file path the planner originally uploaded
    jobs: the list of Job objects after allocation (and any AI-driven edits)
    output_path: where to save the result

    Raises ValueError if the original file's header row doesn't have
    recognizable Vehicle/Driver columns -- better to fail loudly than
    silently write into the wrong cells.
    """
    wb = openpyxl.load_workbook(original_path)  # preserves formatting/styles/formulas
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for idx, header in enumerate(header_row, start=1):  # openpyxl columns are 1-indexed
        key = _normalize_header(header)
        if key in _HEADER_ALIASES:
            col_index[_HEADER_ALIASES[key]] = idx

    if "vehicle" not in col_index or "driver" not in col_index:
        raise ValueError(
            "Could not find 'Vehicle' and 'Driver' columns in this file's header row -- "
            "nothing was changed."
        )

    jobs_by_row = {j.row_number: j for j in jobs}

    for row_number, job in jobs_by_row.items():
        driver_text = ""
        vehicle_text = ""
        if job.assigned_driver_id is not None:
            driver_text = job.assignment_note.replace("In-house: ", "")
            vehicle_text = job.assigned_vehicle_plate
        elif job.assigned_supplier_unit:
            # Matches the real-world convention: supplier name goes in the
            # Driver column, Vehicle column stays blank for hired units.
            driver_text = job.assigned_supplier_unit
        # Unresolved jobs: leave both blank -- makes any gap obvious on the sheet.

        ws.cell(row=row_number, column=col_index["driver"]).value = driver_text
        ws.cell(row=row_number, column=col_index["vehicle"]).value = vehicle_text

    wb.save(output_path)
    return output_path
